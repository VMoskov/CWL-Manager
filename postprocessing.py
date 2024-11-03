from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from argparse import ArgumentParser
import re


coef = {'c1': 1, 'c2': 0.9, 'c3': 0.8, 'm1': 0.7, 'm2': 0.6, 'm3': 0.5}


if __name__ == '__main__':
    parser = ArgumentParser()
    parser.add_argument('--sheet_name', type=str)
    parser.add_argument('--number_of_teams', type=int)
    args = parser.parse_args()

    sheet_name = args.sheet_name
    teams = args.number_of_teams
    file_path = Path('CWL.xlsx')

    df = pd.read_excel(file_path, header=None, sheet_name=sheet_name)
    all_results = pd.DataFrame()

    for i in range(teams):
        team = df.iloc[2:21, i*8+1:i*8+5]
        league = re.search('\((\w+)\)', df.iloc[0, i*8]).group(1)

        team.columns = ['player', 'attacks', 'stars', '%']
        team['score'] = team['stars'] * coef[league]

        team = team.sort_values(['score', '%'], ascending=False)
        all_results = pd.concat([all_results, team])
        
        book = load_workbook(file_path, sheet_name=sheet_name)
        # Write team results
        for row in range(len(team)):
            for col in range(len(team.columns)):
                book[sheet_name].cell(row=row + 3, column=i * 8 + 2 + col, value=team.iloc[row, col])

        book.save(file_path)
        book.close()
    
    # Write final results for the whole clan
    all_results = all_results.sort_values(['score', '%'], ascending=False)
    all_results['rank'] = range(1, len(all_results) + 1)
    all_results = all_results.loc[:, ['rank', 'player', 'stars', 'score']]
    # remove rows with NaN values
    all_results = all_results.dropna()
    for row in range(len(all_results)):
        for col in range(len(all_results.columns)):
            # read value in the cell from the excel file
            book[sheet_name].cell(row=row + 30, column=12 + col, value=all_results.iloc[row, col])
    book.save(file_path)
    book.close()