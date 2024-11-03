from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from argparse import ArgumentParser
import requests, json
import re


hashtag = '%23'  # hashtag (#) needs to be replaced with %23
CROATIA = f'{hashtag}80P90LVY'  
CRO_ELITA_2 = f'{hashtag}2LLCVGU9'
TOKEN = open('auth.txt').read().strip()


def get_players(clan_tag):
    url = f'https://api.clashofclans.com/v1/clans/{clan_tag}/members?limit=50'
    response = requests.get(url, headers={
        'Accept': 'application/json',
        'Authorization': f'Bearer {TOKEN}'
    })
    if response.status_code == 200:
        # extract attributes 'name' and 'townHallLevel' from the response
        players = response.json()['items']
        players = [{'name': player['name'], 'th': player['townHallLevel']} for player in players]
        return players
    else:
        print(f'Error: {response.status_code}, {response.text}')
        return None
    

def get_clan_tag():
    # TODO: extract clan tag from the excel file
    ...
    

if __name__ == '__main__':
    parser = ArgumentParser()
    parser.add_argument('--mode', type=str)  # Choose between "extract" and "distribute"
    parser.add_argument('--sheet_name', type=str)
    parser.add_argument('--number_of_clans', type=int, default=2)  # Number of clans to extract players from
    args = parser.parse_args()

    mode = args.mode
    sheet_name = args.sheet_name
    file_path = Path('CWL.xlsx')
    book = load_workbook(file_path)

    if mode == 'extract':
        df = pd.DataFrame()
        croatia = get_players(CROATIA)
        cro_elita_2 = get_players(CRO_ELITA_2)
        [player.update({'clan': 'CROATIA'}) for player in croatia]
        [player.update({'clan': 'CRO ELITA 2'}) for player in cro_elita_2]
        players = croatia + cro_elita_2
        players = list(map(lambda player: {**player, 'name': player['name'].replace('=', '')}, players))  # there is a player named =dart=

        df = pd.DataFrame(players)
        df = df[['name', 'th', 'clan']]

        for row in range(len(df)):
            for col in range(len(df.columns)):
                book[sheet_name].cell(row=row + 30, column=6 + col, value=df.iloc[row, col])
        
    elif mode == 'distribute':
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        df = df.iloc[28:130, 5:9]
        df.columns = ['name', 'th', 'clan', 'team']

        number_of_teams = df['team'].max()
        for i in range(number_of_teams):
            team = df[df['team'] == i + 1]
            team = team.sort_values('th', ascending=False)
            team = team[['name']]
            for row in range(len(team)):
                for col in range(len(team.columns)):
                    book[sheet_name].cell(row=row + 3, column=i * 8 + 2 + col, value=team.iloc[row, col])
    else:
        print('Invalid mode')

    book.save(file_path)
    book.close()

   